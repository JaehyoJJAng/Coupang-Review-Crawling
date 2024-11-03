from bs4 import BeautifulSoup as bs
from pathlib import Path
from openpyxl import Workbook
from fake_useragent import UserAgent
from requests.exceptions import RequestException
import time
import os
import re
import requests as rq
import json
import math
import sys


def get_headers(key: str) -> dict[str, dict[str, str]] | None:
    """Get Headers"""
    JSON_FILE: str = "json/headers.json"
    with open(JSON_FILE, "r", encoding="UTF-8") as file:
        headers = json.loads(file.read())

    try:
        return headers[key]
    except:
        raise EnvironmentError(f"Set the {key}")


class Coupang:
    @staticmethod
    def get_product_code(url: str) -> str:
        prod_code: str = url.split("products/")[-1].split("?")[0]
        return prod_code

    @staticmethod
    def get_soup_object(resp: rq.Response) -> bs:
        return bs(resp.text, "html.parser")

    def __init__(self) -> None:
        self.__headers: dict[str, str] = get_headers(key="headers")
        self.base_review_url: str = "https://www.coupang.com/vp/product/reviews"
        self.sd = SaveData()
        self.retries = 5
        self.delay = 2

    def get_product_info(self, prod_code: str) -> tuple:
        url = f"https://www.coupang.com/vp/products/{prod_code}"
        resp = rq.get(url=url, headers=self.__headers)
        soup = self.get_soup_object(resp=resp)

        return (
            soup.select_one("h1.prod-buy-header__title").text.strip(),
            int(re.sub("[^0-9]", "", soup.select("span.count")[0].text.strip())),
        )

    def start(self) -> None:
        self.sd.create_directory()
        URL: str = self.input_review_url()
        self.__headers["Referer"] = URL
        prod_code: str = self.get_product_code(url=URL)

        # 상품 정보 추출
        self.title, review_count = self.get_product_info(prod_code=prod_code)
        if review_count > 1500:
            review_pages = 300
        else:
            review_pages: int = self.calculate_total_pages(review_count)

        # Set payload
        payloads = [
            {
                "productId": prod_code,
                "page": page,
                "size": 5,
                "sortBy": "ORDER_SCORE_ASC",
                "ratings": "",
                "q": "",
                "viRoleCode": 2,
                "ratingSummary": True,
            }
            for page in range(1, review_pages + 1)
        ]

        # 데이터 추출
        for payload in payloads:
            self.fetch(payload=payload)

    def fetch(self, payload: list[dict]) -> None:
        now_page: str = payload["page"]
        print(f"\n[INFO] Start crawling page {now_page} ...\n")
        attempt: int = 0
        while attempt < self.retries:
            try:
                with rq.Session() as session:
                    with session.get(
                        url=self.base_review_url,
                        headers=self.__headers,
                        params=payload,
                        timeout=10,
                    ) as resp:
                        resp.raise_for_status()
                        html = resp.text
                        soup = bs(html, "html.parser")

                        # 상품명
                        title = soup.select_one("h1.prod-buy-header__title")
                        if title == None or title.text == "":
                            title = "-"
                        else:
                            title = title.text.strip()

                        # Article Boxes
                        article_lenth = len(
                            soup.select("article.sdp-review__article__list")
                        )

                        for idx in range(article_lenth):
                            dict_data: dict[str, str | int] = dict()
                            articles = soup.select("article.sdp-review__article__list")

                            # 리뷰 날짜
                            review_date = articles[idx].select_one(
                                "div.sdp-review__article__list__info__product-info__reg-date"
                            )
                            if review_date == None or review_date.text == "":
                                review_date = "-"
                            else:
                                review_date = review_date.text.strip()

                            # 구매자 이름
                            user_name = articles[idx].select_one(
                                "span.sdp-review__article__list__info__user__name"
                            )
                            if user_name == None or user_name.text == "":
                                user_name = "-"
                            else:
                                user_name = user_name.text.strip()

                            # 평점
                            rating = articles[idx].select_one(
                                "div.sdp-review__article__list__info__product-info__star-orange"
                            )
                            if rating == None:
                                rating = 0
                            else:
                                rating = int(rating.attrs["data-rating"])

                            # 구매자 상품명
                            prod_name = articles[idx].select_one(
                                "div.sdp-review__article__list__info__product-info__name"
                            )
                            if prod_name == None or prod_name.text == "":
                                prod_name = "-"
                            else:
                                prod_name = prod_name.text.strip()

                            # 헤드라인(타이틀)
                            headline = articles[idx].select_one(
                                "div.sdp-review__article__list__headline"
                            )
                            if headline == None or headline.text == "":
                                headline = "등록된 헤드라인이 없습니다"
                            else:
                                headline = headline.text.strip()

                            # 리뷰 내용
                            review_content = articles[idx].select_one(
                                "div.sdp-review__article__list__review > div"
                            )
                            if review_content == None:
                                review_content = "등록된 리뷰내용이 없습니다"
                            else:
                                review_content = re.sub(
                                    "[\n\t]", "", review_content.text.strip()
                                )

                            # 맛 만족도
                            answer = articles[idx].select_one(
                                "span.sdp-review__article__list__survey__row__answer"
                            )
                            if answer == None or answer.text == "":
                                answer = "맛 평가 없음"
                            else:
                                answer = answer.text.strip()

                            dict_data["title"] = self.title
                            dict_data["prod_name"] = prod_name
                            dict_data["review_date"] = review_date
                            dict_data["user_name"] = user_name
                            dict_data["rating"] = rating
                            dict_data["headline"] = headline
                            dict_data["review_content"] = review_content
                            dict_data["answer"] = answer
                            self.sd.save(datas=dict_data)
                            print(dict_data, "\n")
                        time.sleep(1)
                        return
            except RequestException as e:
                attempt += 1
                print(f"Attempt {attempt}/{self.retries} failed: {e}")
                if attempt < self.retries:
                    time.sleep(self.delay)
                else:
                    print(f"최대 요청 만료! 다시 실행하세요.")
                    sys.exit()

    @staticmethod
    def clear_console() -> None:
        command: str = "clear"
        if os.name in ("nt", "dos"):
            command = "cls"
        os.system(command=command)

    def input_review_url(self) -> str:
        while True:
            self.clear_console()
            review_url: str = input(
                "원하시는 상품의 URL 주소를 입력해주세요\n\nEx)\nhttps://www.coupang.com/vp/products/7335597976?itemId=18741704367&vendorItemId=85873964906&q=%ED%9E%98%EB%82%B4%EB%B0%94+%EC%B4%88%EC%BD%94+%EC%8A%A4%EB%8B%88%EC%BB%A4%EC%A6%88&itemsCount=36&searchId=0c5c84d537bc41d1885266961d853179&rank=2&isAddedCart=\n\n:"
            )
            if not review_url:
                # 터미널 초기화
                self.clear_console()

                print("URL 주소가 입력되지 않았습니다")
                continue
            return review_url

    def calculate_total_pages(self, review_counts: int) -> int:
        reviews_per_page: int = 5
        return int(math.ceil(review_counts / reviews_per_page))


class SaveData:
    def __init__(self) -> None:
        self.wb: Workbook = Workbook()
        self.ws = self.wb.active
        self.ws.append(["이름", "작성일자", "평점", "리뷰 내용", "맛 만족도"])
        self.row: int = 2
        self.dir_name: str = "Coupang-reviews"
        self.create_directory()

    def create_directory(self) -> None:
        if not os.path.exists(self.dir_name):
            os.makedirs(self.dir_name)

    def save(self, datas: dict[str, str | int]) -> None:
        file_name: str = os.path.join(self.dir_name, datas["title"] + ".xlsx")
        self.ws[f"A{self.row}"] = datas["user_name"]
        self.ws[f"B{self.row}"] = datas["review_date"]
        self.ws[f"C{self.row}"] = datas["rating"]
        self.ws[f"D{self.row}"] = datas["review_content"]
        self.ws[f"E{self.row}"] = datas["answer"]
        self.row += 1
        self.wb.save(filename=file_name)

    def __del__(self) -> None:
        self.wb.close()
